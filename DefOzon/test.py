import openpyxl


def test():
    # Загрузка файла Excel
    file_path = rf'C:\Users\Dimulka\PycharmProjects\NewVerUpdateSites\UpdateResultYandex\ya1.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    # Выбор активного листа (может потребоваться его изменение в зависимости от вашей структуры)
    sheet = workbook.active

    # Проход по всем столбцам
    for column_index in range(1, sheet.max_column + 1):
        # Проход по всем строкам в текущем столбце
        for row_index in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=column_index)

            # Проверка наличия символа ₽ в ячейке (преобразование в строку перед проверкой)
            if '₽' in str(cell.value):
                # Удаление символов после ₽
                cell.value = str(cell.value).split('₽')[0]

            # Проверка наличия значения 0 в ячейке
            if cell.value == 0:
                # Очистка ячейки со значением 0
                cell.value = None

                # Очистка ячейки в следующем столбце (если она существует)
                if column_index < sheet.max_column:
                    next_cell = sheet.cell(row=row_index, column=column_index + 1)
                    next_cell.value = None

    # Сохранение изменений
    workbook.save('Yandex1.xlsx')


# Вызов функции
# test()
